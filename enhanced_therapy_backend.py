"""
Enhanced Therapeutic Companion Web Backend
Supports complete weekly tracking with Excel export
Deployment-ready with security features and multi-user support
"""

from flask import Flask, request, jsonify, send_file, Response, session
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from functools import wraps
import json
import os
import io
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import hashlib
import secrets
import shutil


print("=" * 50)
print("CHECKING ENVIRONMENT VARIABLES AT STARTUP:")
print(f"SYSTEM_EMAIL: {os.environ.get('SYSTEM_EMAIL')}")
print(f"SYSTEM_EMAIL_PASSWORD: {os.environ.get('SYSTEM_EMAIL_PASSWORD')}")
print("=" * 50)

# Also check if .env file exists
if os.path.exists('.env'):
    print(".env file found in project directory!")
    with open('.env', 'r') as f:
        print("Contents:")
        print(f.read())
print("=" * 50)


# Optional: Load environment variables from .env file for local development
try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    # dotenv not installed - environment variables must be set another way
    pass

# Optional: SendGrid support
try:
    import sendgrid
    from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
    import base64

    SENDGRID_AVAILABLE = True
except ImportError:
    SENDGRID_AVAILABLE = False

# Import the social worker components
from socialworkcountry import GlobalSocialWorkerChatbot, PatientProfile

# Create Flask app
app = Flask(__name__)

# Configure CORS for production
if os.environ.get('PRODUCTION'):
    CORS(app, origins=[os.environ.get('ALLOWED_ORIGINS', '*')])
else:
    CORS(app)  # Allow all origins in development

# Configure session security
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
if os.environ.get('PRODUCTION'):
    app.config['SESSION_COOKIE_SECURE'] = True  # HTTPS only
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Initialize rate limiter
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"]
)

# Create data directories
os.makedirs('therapy_data', exist_ok=True)
os.makedirs('therapy_data/patients', exist_ok=True)
os.makedirs('therapy_data/checkins', exist_ok=True)
os.makedirs('therapy_data/reports', exist_ok=True)
os.makedirs('therapy_data/excel_exports', exist_ok=True)
os.makedirs('therapy_data/therapists', exist_ok=True)
os.makedirs('therapy_data/logs', exist_ok=True)

# Initialize the social worker chatbot
chatbot = GlobalSocialWorkerChatbot()


# ============= AUTHENTICATION SYSTEM =============

def generate_access_token():
    """Generate secure access token"""
    return secrets.token_urlsafe(32)


def hash_password(password):
    """Hash a password for storing"""
    return hashlib.sha256(password.encode()).hexdigest()


def validate_therapist_token(token):
    """Validate a therapist's access token"""
    therapists_dir = os.path.join('therapy_data', 'therapists')
    if not os.path.exists(therapists_dir):
        return None

    for filename in os.listdir(therapists_dir):
        if filename.endswith('.json'):
            with open(os.path.join(therapists_dir, filename), 'r') as f:
                therapist = json.load(f)
                if therapist.get('access_token') == token and therapist.get('active', True):
                    return therapist

    return None


def require_auth(f):
    """Decorator to require authentication"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check for master token (for development/admin)
        master_token = os.environ.get('MASTER_TOKEN')

        auth_header = request.headers.get('Authorization', '')
        token = auth_header.replace('Bearer ', '')

        # Allow master token
        if master_token and token == master_token:
            request.therapist = {'email': 'admin@system', 'name': 'System Admin'}
            return f(*args, **kwargs)

        # Validate regular token
        therapist = validate_therapist_token(token)
        if not therapist:
            return jsonify({'error': 'Invalid or expired token'}), 401

        # Add therapist info to request
        request.therapist = therapist
        return f(*args, **kwargs)

    return decorated_function


def mock_auth(f):
    """Decorator to add mock therapist for development"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        request.therapist = {
            'email': 'admin@system',
            'name': 'System Admin',
            'organization': 'Development'
        }
        return f(*args, **kwargs)

    return decorated_function


# ============= PUBLIC ENDPOINTS =============

@app.route('/')
def index():
    """Serve the main HTML file"""
    # First try client.html (your file)
    if os.path.exists('client.html'):
        with open('client.html', 'r', encoding='utf-8') as f:
            return f.read()
    # Fallback to therapy_tracker.html
    elif os.path.exists('therapy_tracker.html'):
        with open('therapy_tracker.html', 'r', encoding='utf-8') as f:
            return f.read()
    else:
        return """
        <html>
        <body>
            <h1>Enhanced Therapeutic Companion Server Running</h1>
            <p>Please ensure client.html or therapy_tracker.html is in the same directory as this script.</p>
        </body>
        </html>
        """


@app.route('/api/therapy/register-therapist', methods=['POST'])
@limiter.limit("5 per day")  # Prevent registration abuse
def register_therapist():
    """Register a new therapist"""
    try:
        data = request.json

        # Validate input
        required_fields = ['email', 'name', 'organization', 'password']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing {field}'}), 400

        # Check if therapist already exists
        therapist_file = os.path.join('therapy_data', 'therapists', f"{data['email']}.json")
        if os.path.exists(therapist_file):
            return jsonify({'success': False, 'error': 'Email already registered'}), 400

        # Generate access token
        access_token = generate_access_token()

        # Save therapist data
        therapist_data = {
            'email': data['email'],
            'name': data['name'],
            'organization': data['organization'],
            'password_hash': hash_password(data['password']),
            'access_token': access_token,
            'created_at': datetime.now().isoformat(),
            'active': True
        }

        os.makedirs(os.path.dirname(therapist_file), exist_ok=True)
        with open(therapist_file, 'w') as f:
            json.dump(therapist_data, f, indent=2)

        # Log registration
        log_activity('therapist_registration', {'email': data['email']})

        return jsonify({
            'success': True,
            'message': 'Registration successful',
            'access_token': access_token,
            'therapist': {
                'email': data['email'],
                'name': data['name'],
                'organization': data['organization']
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/therapy/login', methods=['POST'])
@limiter.limit("10 per hour")
def login_therapist():
    """Login for therapists"""
    try:
        data = request.json
        email = data.get('email')
        password = data.get('password')

        if not email or not password:
            return jsonify({'success': False, 'error': 'Missing email or password'}), 400

        # Load therapist data
        therapist_file = os.path.join('therapy_data', 'therapists', f"{email}.json")
        if not os.path.exists(therapist_file):
            return jsonify({'success': False, 'error': 'Invalid credentials'}), 401

        with open(therapist_file, 'r') as f:
            therapist = json.load(f)

        # Verify password
        if therapist['password_hash'] != hash_password(password):
            return jsonify({'success': False, 'error': 'Invalid credentials'}), 401

        # Check if account is active
        if not therapist.get('active', True):
            return jsonify({'success': False, 'error': 'Account deactivated'}), 401

        # Generate new token for this session
        new_token = generate_access_token()
        therapist['access_token'] = new_token
        therapist['last_login'] = datetime.now().isoformat()

        # Save updated data
        with open(therapist_file, 'w') as f:
            json.dump(therapist, f, indent=2)

        # Log login
        log_activity('therapist_login', {'email': email})

        return jsonify({
            'success': True,
            'access_token': new_token,
            'therapist': {
                'email': therapist['email'],
                'name': therapist['name'],
                'organization': therapist['organization']
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============= PROTECTED THERAPY TRACKING ENDPOINTS =============

@app.route('/api/therapy/save-patient', methods=['POST'])
@mock_auth  # Using mock_auth instead of require_auth for development
def save_therapy_patient():
    """Save therapy patient enrollment data"""
    try:
        data = request.json
        patient_id = data.get('patientId')
        patient_data = data.get('patientData')

        if not patient_id or not patient_data:
            return jsonify({
                'success': False,
                'error': 'Missing patient ID or data'
            }), 400

        # Add enrollment metadata
        patient_data['enrollmentTimestamp'] = datetime.now().isoformat()
        patient_data['enrolledBy'] = request.therapist['email']
        patient_data['therapistOrganization'] = request.therapist.get('organization', '')
        patient_data['weeklyReports'] = []

        # Save patient data
        filename = f'patient_{patient_id}.json'
        filepath = os.path.join('therapy_data', 'patients', filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(patient_data, f, indent=2, ensure_ascii=False)

        # Log activity
        log_activity('patient_enrolled', {
            'patient_id': patient_id,
            'therapist': request.therapist['email']
        })

        return jsonify({
            'success': True,
            'message': 'Patient enrolled successfully',
            'filename': filename
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/therapy/save-checkin', methods=['POST'])
@mock_auth  # Using mock_auth instead of require_auth for development
def save_therapy_checkin():
    """Save comprehensive daily check-in data"""
    try:
        data = request.json
        patient_id = data.get('patientId')
        checkin_data = data.get('checkinData')

        if not patient_id or not checkin_data:
            return jsonify({
                'success': False,
                'error': 'Missing patient ID or check-in data'
            }), 400

        # Verify patient belongs to therapist's organization
        patient_file = os.path.join('therapy_data', 'patients', f'patient_{patient_id}.json')
        if not os.path.exists(patient_file):
            return jsonify({'success': False, 'error': 'Patient not found'}), 404

        with open(patient_file, 'r') as f:
            patient = json.load(f)

        # Check authorization
        if patient.get('enrolledBy') != request.therapist['email'] and request.therapist['email'] != 'admin@system':
            return jsonify({'success': False, 'error': 'Unauthorized access to patient'}), 403

        # Validate check-in data
        required_fields = ['date', 'time', 'emotional', 'medication', 'activity']
        for field in required_fields:
            if field not in checkin_data:
                return jsonify({
                    'success': False,
                    'error': f'Missing required field: {field}'
                }), 400

        # Add metadata
        checkin_data['serverTimestamp'] = datetime.now().isoformat()
        checkin_data['recordedBy'] = request.therapist['email']

        # Create patient checkin directory
        patient_dir = os.path.join('therapy_data', 'checkins', patient_id)
        os.makedirs(patient_dir, exist_ok=True)

        # Save check-in data
        date = checkin_data.get('date')
        filename = f'checkin_{date}.json'
        filepath = os.path.join(patient_dir, filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(checkin_data, f, indent=2, ensure_ascii=False)

        # Log activity
        log_activity('checkin_recorded', {
            'patient_id': patient_id,
            'date': date,
            'therapist': request.therapist['email']
        })

        return jsonify({
            'success': True,
            'message': 'Daily check-in saved successfully',
            'filename': filename
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/therapy/get-week-data/<patient_id>/<week>', methods=['GET'])
@mock_auth  # Using mock_auth instead of require_auth for development
def get_week_data(patient_id, week):
    """Get all check-in data for a specific week"""
    try:
        # Verify authorization
        patient_file = os.path.join('therapy_data', 'patients', f'patient_{patient_id}.json')
        if os.path.exists(patient_file):
            with open(patient_file, 'r') as f:
                patient = json.load(f)

            if patient.get('enrolledBy') != request.therapist['email'] and request.therapist['email'] != 'admin@system':
                return jsonify({'success': False, 'error': 'Unauthorized access'}), 403

        week_data = {}
        checkin_dir = os.path.join('therapy_data', 'checkins', patient_id)

        if os.path.exists(checkin_dir):
            # Parse week string
            year, week_num = week.split('-W')
            year = int(year)
            week_num = int(week_num)

            # Calculate week dates
            jan_4 = datetime(year, 1, 4)
            week_1_monday = jan_4 - timedelta(days=jan_4.weekday())
            week_start = week_1_monday + timedelta(weeks=week_num - 1)

            # Get data for each day
            for i in range(7):
                date = week_start + timedelta(days=i)
                date_str = date.strftime('%Y-%m-%d')

                checkin_file = os.path.join(checkin_dir, f'checkin_{date_str}.json')
                if os.path.exists(checkin_file):
                    with open(checkin_file, 'r', encoding='utf-8') as f:
                        week_data[date_str] = json.load(f)

        return jsonify({
            'success': True,
            'weekData': week_data
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/therapy/get-all-patients', methods=['GET'])
@mock_auth  # Using mock_auth instead of require_auth for development
def get_all_therapy_patients():
    """Get list of all enrolled therapy patients for this therapist"""
    try:
        patients = []
        patients_dir = os.path.join('therapy_data', 'patients')

        if os.path.exists(patients_dir):
            for filename in os.listdir(patients_dir):
                if filename.startswith('patient_') and filename.endswith('.json'):
                    filepath = os.path.join(patients_dir, filename)
                    with open(filepath, 'r', encoding='utf-8') as f:
                        patient_data = json.load(f)

                        # Only show patients enrolled by this therapist
                        if (patient_data.get('enrolledBy') == request.therapist['email'] or
                                request.therapist['email'] == 'admin@system'):
                            patients.append(patient_data)

        return jsonify({
            'success': True,
            'patients': patients
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/therapy/generate-excel-report/<patient_id>/<week>', methods=['GET'])
@mock_auth  # Using mock_auth instead of require_auth for development
@limiter.limit("20 per hour")
def generate_excel_report(patient_id, week):
    """Generate comprehensive Excel report for a patient's week"""
    try:
        # Get patient data and verify authorization
        patient_file = os.path.join('therapy_data', 'patients', f'patient_{patient_id}.json')
        if not os.path.exists(patient_file):
            return jsonify({'success': False, 'error': 'Patient not found'}), 404

        with open(patient_file, 'r', encoding='utf-8') as f:
            patient_data = json.load(f)

        if patient_data.get('enrolledBy') != request.therapist['email'] and request.therapist[
            'email'] != 'admin@system':
            return jsonify({'success': False, 'error': 'Unauthorized access'}), 403

        # Get week data
        week_response = get_week_data(patient_id, week)
        week_data_json = week_response.get_json()
        week_data = week_data_json.get('weekData', {})

        # Create Excel workbook
        wb = openpyxl.Workbook()

        # Create Summary Sheet
        summary_sheet = wb.active
        summary_sheet.title = "Weekly Summary"

        # Styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add patient information
        summary_sheet['A1'] = "WEEKLY THERAPY TRACKING REPORT"
        summary_sheet['A1'].font = Font(bold=True, size=16)
        summary_sheet.merge_cells('A1:F1')

        summary_sheet['A3'] = "Patient Information"
        summary_sheet['A3'].font = subheader_font
        summary_sheet.merge_cells('A3:B3')

        patient_info_rows = [
            ("Patient ID:", patient_data['patientId']),
            ("Patient Name:", patient_data['name']),
            ("Therapist:", patient_data['therapistName']),
            ("Therapist Email:", patient_data['therapistEmail']),
            ("Week:", week),
            ("Report Generated:", datetime.now().strftime("%Y-%m-%d %H:%M"))
        ]

        row = 4
        for label, value in patient_info_rows:
            summary_sheet[f'A{row}'] = label
            summary_sheet[f'B{row}'] = value
            summary_sheet[f'A{row}'].font = Font(bold=True)
            row += 1

        # Add summary statistics
        summary_sheet['D3'] = "Weekly Statistics"
        summary_sheet['D3'].font = subheader_font
        summary_sheet.merge_cells('D3:F3')

        # Calculate statistics
        total_days = 7
        completed_days = len(week_data)

        if completed_days > 0:
            total_emotional = sum(data['emotional']['value'] for data in week_data.values())
            total_medication = sum(data['medication']['value'] for data in week_data.values())
            total_activity = sum(data['activity']['value'] for data in week_data.values())

            avg_emotional = total_emotional / completed_days
            avg_medication = total_medication / completed_days
            avg_activity = total_activity / completed_days
        else:
            avg_emotional = avg_medication = avg_activity = 0

        stats_rows = [
            ("Completion Rate:", f"{completed_days}/{total_days} ({completed_days / 7 * 100:.1f}%)"),
            ("Avg Emotional State:", f"{avg_emotional:.2f}/5" if completed_days > 0 else "N/A"),
            ("Avg Medication Adherence:", f"{avg_medication:.2f}/5" if completed_days > 0 else "N/A"),
            ("Avg Physical Activity:", f"{avg_activity:.2f}/5" if completed_days > 0 else "N/A")
        ]

        row = 4
        for label, value in stats_rows:
            summary_sheet[f'D{row}'] = label
            summary_sheet[f'E{row}'] = value
            summary_sheet[f'D{row}'].font = Font(bold=True)
            row += 1

        # Create Daily Data Sheet
        daily_sheet = wb.create_sheet("Daily Check-ins")

        # Headers for daily data
        headers = ["Date", "Day", "Time", "Emotional State", "Emotional Notes",
                   "Medication Adherence", "Medication Notes", "Physical Activity",
                   "Activity Notes", "Check-in Status"]

        for col, header in enumerate(headers, 1):
            cell = daily_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Parse week to get dates
        year, week_num = week.split('-W')
        year = int(year)
        week_num = int(week_num)

        # Calculate week start date
        jan_4 = datetime(year, 1, 4)
        week_1_monday = jan_4 - timedelta(days=jan_4.weekday())
        week_start = week_1_monday + timedelta(weeks=week_num - 1)

        # Add daily data
        days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

        for day_num in range(7):
            current_date = week_start + timedelta(days=day_num)
            date_str = current_date.strftime('%Y-%m-%d')

            row = day_num + 2
            daily_sheet.cell(row=row, column=1, value=date_str).border = border
            daily_sheet.cell(row=row, column=2, value=days_of_week[day_num]).border = border

            if date_str in week_data:
                data = week_data[date_str]
                daily_sheet.cell(row=row, column=3, value=data.get('time', '')).border = border
                daily_sheet.cell(row=row, column=4, value=data['emotional']['value']).border = border
                daily_sheet.cell(row=row, column=5, value=data['emotional'].get('notes', '')).border = border

                # Medication value with text labels
                med_value = data['medication']['value']
                medication_text = {
                    0: "Not Applicable",
                    1: "No Doses",
                    3: "Partial Doses",
                    5: "Yes, All Doses"
                }.get(med_value, str(med_value))
                daily_sheet.cell(row=row, column=6, value=medication_text).border = border
                daily_sheet.cell(row=row, column=7, value=data['medication'].get('notes', '')).border = border
                daily_sheet.cell(row=row, column=8, value=data['activity']['value']).border = border
                daily_sheet.cell(row=row, column=9, value=data['activity'].get('notes', '')).border = border
                daily_sheet.cell(row=row, column=10, value="Completed").border = border

                # Color code emotional state
                emotional_cell = daily_sheet.cell(row=row, column=4)
                emotional_value = emotional_cell.value
                if emotional_value >= 4:
                    emotional_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif emotional_value == 3:
                    emotional_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    emotional_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # Color code medication adherence
                medication_cell = daily_sheet.cell(row=row, column=6)
                if med_value == 0:  # Not Applicable - no color
                    pass
                elif med_value == 1:  # No Doses - red
                    medication_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif med_value == 3:  # Partial Doses - yellow
                    medication_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif med_value == 5:  # Yes, All Doses - green
                    medication_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                # Color code physical activity
                activity_cell = daily_sheet.cell(row=row, column=8)
                activity_value = activity_cell.value
                if activity_value >= 4:
                    activity_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif activity_value == 3:
                    activity_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    activity_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:
                daily_sheet.cell(row=row, column=3, value="-").border = border
                for col in range(4, 10):
                    daily_sheet.cell(row=row, column=col, value="-").border = border
                cell = daily_sheet.cell(row=row, column=10, value="No Response")
                cell.border = border
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Create Detailed Notes Sheet
        notes_sheet = wb.create_sheet("Detailed Notes")

        # Headers for notes
        notes_headers = ["Date", "Category", "Rating", "Notes"]
        for col, header in enumerate(notes_headers, 1):
            cell = notes_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Add all notes
        row = 2
        for date_str in sorted(week_data.keys()):
            data = week_data[date_str]

            # Emotional notes
            if data['emotional'].get('notes'):
                notes_sheet.cell(row=row, column=1, value=date_str).border = border
                notes_sheet.cell(row=row, column=2, value="Emotional").border = border
                notes_sheet.cell(row=row, column=3, value=data['emotional']['value']).border = border
                notes_sheet.cell(row=row, column=4, value=data['emotional']['notes']).border = border
                row += 1

            # Medication notes
            if data['medication'].get('notes'):
                notes_sheet.cell(row=row, column=1, value=date_str).border = border
                notes_sheet.cell(row=row, column=2, value="Medication").border = border
                med_value = data['medication']['value']
                medication_text = {
                    0: "Not Applicable",
                    1: "No Doses",
                    3: "Partial Doses",
                    5: "Yes, All Doses"
                }.get(med_value, str(med_value))
                notes_sheet.cell(row=row, column=3, value=medication_text).border = border
                notes_sheet.cell(row=row, column=4, value=data['medication']['notes']).border = border
                row += 1

            # Activity notes
            if data['activity'].get('notes'):
                notes_sheet.cell(row=row, column=1, value=date_str).border = border
                notes_sheet.cell(row=row, column=2, value="Physical Activity").border = border
                notes_sheet.cell(row=row, column=3, value=data['activity']['value']).border = border
                notes_sheet.cell(row=row, column=4, value=data['activity']['notes']).border = border
                row += 1

        # Auto-adjust column widths
        for sheet in [summary_sheet, daily_sheet, notes_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Save Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"therapy_report_{patient_id}_{week}_{timestamp}.xlsx"
        filepath = os.path.join('therapy_data', 'excel_exports', filename)

        wb.save(filepath)

        # Log activity
        log_activity('report_generated', {
            'patient_id': patient_id,
            'week': week,
            'therapist': request.therapist['email']
        })

        # Return file as download
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/therapy/email-report', methods=['POST'])
@mock_auth  # Using mock_auth instead of require_auth for development
@limiter.limit("10 per hour")
def email_therapy_report():
    """Generate Excel report and send via email using system email account"""
    try:
        print(f"DEBUG: SYSTEM_EMAIL env var: {os.environ.get('SYSTEM_EMAIL')}")
        print(f"DEBUG: SYSTEM_EMAIL_PASSWORD env var: {os.environ.get('SYSTEM_EMAIL_PASSWORD')}")
        data = request.json
        patient_id = data.get('patientId')
        week = data.get('week')
        custom_recipient = data.get('customRecipient')  # Optional custom recipient

        # Get patient data and verify authorization
        patient_file = os.path.join('therapy_data', 'patients', f'patient_{patient_id}.json')
        if not os.path.exists(patient_file):
            return jsonify({'success': False, 'error': 'Patient not found'}), 404

        with open(patient_file, 'r', encoding='utf-8') as f:
            patient_data = json.load(f)

        if patient_data.get('enrolledBy') != request.therapist['email'] and request.therapist[
            'email'] != 'admin@system':
            return jsonify({'success': False, 'error': 'Unauthorized access'}), 403

        # Determine recipient
        recipient_email = custom_recipient if custom_recipient else patient_data['therapistEmail']

        # First, generate the Excel report
        excel_response = generate_excel_report(patient_id, week)

        if not isinstance(excel_response, Response):
            return jsonify({'success': False, 'error': 'Failed to generate Excel report'}), 500

        # Get week data for email content
        week_response = get_week_data(patient_id, week)
        week_data_json = week_response.get_json()
        week_data = week_data_json.get('weekData', {})

        # Calculate summary
        completed_days = len(week_data)
        if completed_days > 0:
            avg_emotional = sum(d['emotional']['value'] for d in week_data.values()) / completed_days

            # Handle medication values properly
            med_values = []
            for d in week_data.values():
                med_val = d['medication']['value']
                if med_val > 0:  # Exclude "Not Applicable"
                    med_values.append(med_val)

            avg_medication = sum(med_values) / len(med_values) if med_values else 0
            avg_activity = sum(d['activity']['value'] for d in week_data.values()) / completed_days
        else:
            avg_emotional = avg_medication = avg_activity = 0

        # Find the Excel file
        excel_files = []
        excel_dir = os.path.join('therapy_data', 'excel_exports')
        if os.path.exists(excel_dir):
            for filename in os.listdir(excel_dir):
                if filename.startswith(f"therapy_report_{patient_id}_{week}_"):
                    excel_files.append(os.path.join(excel_dir, filename))

        if not excel_files:
            return jsonify({
                'success': False,
                'error': 'No Excel report found. Please generate one first.'
            }), 404

        excel_filepath = max(excel_files, key=os.path.getctime)
        excel_filename = os.path.basename(excel_filepath)

        # Prepare email content
        system_name = os.environ.get('SYSTEM_NAME', 'Therapeutic Companion System')
        email_content = f"""
Dear {patient_data['therapistName']},

This is the weekly therapy tracking report for {patient_data['name']} (ID: {patient_id}).

Week: {week}
Completion Rate: {completed_days}/7 days ({completed_days / 7 * 100:.1f}%)

Summary Statistics:
- Average Emotional State: {avg_emotional:.2f}/5
- Average Medication Adherence: {avg_medication:.2f}/5 {"(excluding N/A)" if avg_medication > 0 else ""}
- Average Physical Activity: {avg_activity:.2f}/5

Please find the detailed Excel report attached.

Best regards,
{system_name}

---
This is an automated report. Please do not reply to this email.
Generated by: {request.therapist['name']} ({request.therapist['email']})
Organization: {request.therapist.get('organization', 'N/A')}
        """

        # Try to send email
        email_sent = False
        error_message = None

        # Check for SendGrid first
        if SENDGRID_AVAILABLE and os.environ.get('SENDGRID_API_KEY'):
            try:
                email_sent = send_email_via_sendgrid(
                    recipient_email,
                    f'Weekly Therapy Report - {patient_data["name"]} - Week {week}',
                    email_content,
                    excel_filepath,
                    reply_to=request.therapist['email']
                )
            except Exception as e:
                error_message = str(e)

        # Fall back to SMTP
        if not email_sent:
            system_email_config = get_system_email_config()

            if system_email_config:
                print(f"DEBUG: SYSTEM_EMAIL env var: {os.environ.get('SYSTEM_EMAIL')}")
                print(f"DEBUG: SYSTEM_EMAIL_PASSWORD env var: {os.environ.get('SYSTEM_EMAIL_PASSWORD')}")
                try:
                    # Create message
                    msg = MIMEMultipart()
                    msg['From'] = f"{system_name} <{system_email_config['sender_email']}>"
                    msg['To'] = recipient_email
                    msg['Subject'] = f'Weekly Therapy Report - {patient_data["name"]} - Week {week}'
                    msg['Reply-To'] = request.therapist['email']

                    # Add body
                    msg.attach(MIMEText(email_content, 'plain'))

                    # Add Excel attachment
                    with open(excel_filepath, 'rb') as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename= {excel_filename}'
                        )
                        msg.attach(part)

                    # Send email with better error handling
                    print(f"Attempting to send email from {system_email_config['sender_email']} to {recipient_email}")
                    server = smtplib.SMTP(system_email_config['smtp_server'], system_email_config['smtp_port'])
                    server.set_debuglevel(1)  # Enable debug output
                    server.starttls()
                    server.login(system_email_config['sender_email'], system_email_config['sender_password'])
                    server.send_message(msg)
                    server.quit()

                    email_sent = True
                    print("Email sent successfully!")

                except smtplib.SMTPAuthenticationError as e:
                    error_message = f"Authentication failed: {str(e)}. Please check your email and app password."
                    print(f"SMTP Auth Error: {error_message}")
                except smtplib.SMTPException as e:
                    error_message = f"SMTP error: {str(e)}"
                    print(f"SMTP Error: {error_message}")
                except Exception as e:
                    error_message = f"General error: {str(e)}"
                    print(f"General Error: {error_message}")

        if email_sent:
            # Log email sent
            log_activity('email_sent', {
                'patient_id': patient_id,
                'recipient': recipient_email,
                'week': week,
                'therapist': request.therapist['email']
            })

            return jsonify({
                'success': True,
                'message': 'Email sent successfully',
                'recipient': recipient_email,
                'subject': f'Weekly Therapy Report - {patient_data["name"]} - Week {week}',
                'note': 'Email sent with Excel attachment'
            })
        else:
            # Email not sent - provide preview with actual error
            response_data = {
                'success': True,
                'message': 'Email report prepared',
                'recipient': recipient_email,
                'subject': f'Weekly Therapy Report - {patient_data["name"]} - Week {week}',
                'content': email_content,
                'attachment': excel_filename,
                'attachment_path': excel_filepath
            }

            if system_email_config:
                # Configuration exists but email failed
                response_data['note'] = f'Email configuration found but sending failed: {error_message}'
                response_data['troubleshooting'] = [
                    'Verify your Gmail App Password is correct',
                    'Ensure 2-Factor Authentication is enabled on your Gmail account',
                    'Check that the app password has not expired',
                    'Try generating a new App Password at https://myaccount.google.com/apppasswords'
                ]
            else:
                # No configuration found
                response_data['note'] = 'Email not sent - system email not configured. Contact administrator.'
                response_data['config_example'] = {
                    'sender_email': 'your-email@gmail.com',
                    'sender_password': 'your-app-password',
                    'smtp_server': 'smtp.gmail.com',
                    'smtp_port': 587
                }

            if error_message:
                response_data['error'] = error_message

            return jsonify(response_data)

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============= HELPER FUNCTIONS =============

def get_system_email_config():
    """Get system email configuration"""
    # Priority: Environment variables > Config file > Default

    # Check environment variables
    if os.environ.get('SYSTEM_EMAIL') and os.environ.get('SYSTEM_EMAIL_PASSWORD'):
        return {
            'sender_email': os.environ.get('SYSTEM_EMAIL'),
            'sender_password': os.environ.get('SYSTEM_EMAIL_PASSWORD'),
            'smtp_server': os.environ.get('SMTP_SERVER', 'smtp.gmail.com'),
            'smtp_port': int(os.environ.get('SMTP_PORT', '587'))
        }

    # Check config file
    email_config_file = os.path.join('therapy_data', 'email_config.json')
    if os.path.exists(email_config_file):
        with open(email_config_file, 'r') as f:
            #return json.load(f)
            config = json.load(f)
            print(
                f"DEBUG: Email config loaded - Email: {config.get('sender_email')}, Password length: {len(config.get('sender_password', ''))}")
            return config

    return None


def send_email_via_sendgrid(recipient_email, subject, content, attachment_path, reply_to=None):
    """Send email using SendGrid API"""
    if not SENDGRID_AVAILABLE:
        raise Exception("SendGrid not installed")

    sg_api_key = os.environ.get('SENDGRID_API_KEY')
    if not sg_api_key:
        raise Exception("SendGrid API key not configured")

    # Create message
    from_email = (
        os.environ.get('SYSTEM_EMAIL', 'noreply@therapeutic-companion.com'),
        os.environ.get('SYSTEM_NAME', 'Therapeutic Companion System')
    )

    message = Mail(
        from_email=from_email,
        to_emails=recipient_email,
        subject=subject,
        plain_text_content=content
    )

    if reply_to:
        message.reply_to = reply_to

    # Add attachment
    with open(attachment_path, 'rb') as f:
        data = f.read()
        encoded = base64.b64encode(data).decode()

    attachment = Attachment()
    attachment.file_content = FileContent(encoded)
    attachment.file_type = FileType('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment.file_name = FileName(os.path.basename(attachment_path))
    attachment.disposition = Disposition('attachment')

    message.attachment = attachment

    # Send
    try:
        sg = sendgrid.SendGridAPIClient(api_key=sg_api_key)
        response = sg.send(message)
        return response.status_code == 202
    except Exception as e:
        raise Exception(f"SendGrid error: {str(e)}")


def log_activity(activity_type, data):
    """Log system activity"""
    log_entry = {
        'timestamp': datetime.now().isoformat(),
        'activity': activity_type,
        'data': data
    }

    # Create log file for today
    log_date = datetime.now().strftime('%Y-%m-%d')
    log_file = os.path.join('therapy_data', 'logs', f'activity_{log_date}.json')

    # Read existing log
    if os.path.exists(log_file):
        with open(log_file, 'r') as f:
            log_data = json.load(f)
    else:
        log_data = []

    # Append new entry
    log_data.append(log_entry)

    # Save log
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    with open(log_file, 'w') as f:
        json.dump(log_data, f, indent=2)


# ============= DATA PRIVACY ENDPOINTS =============

@app.route('/api/privacy/delete-patient/<patient_id>', methods=['DELETE'])
@mock_auth  # Using mock_auth instead of require_auth for development
def delete_patient_data(patient_id):
    """GDPR compliance - delete all patient data"""
    try:
        # Verify ownership
        patient_file = os.path.join('therapy_data', 'patients', f'patient_{patient_id}.json')
        if os.path.exists(patient_file):
            with open(patient_file, 'r') as f:
                patient = json.load(f)

            if patient.get('enrolledBy') != request.therapist['email'] and request.therapist['email'] != 'admin@system':
                return jsonify({'success': False, 'error': 'Unauthorized'}), 403

        # Delete patient file
        if os.path.exists(patient_file):
            os.remove(patient_file)

        # Delete all check-ins
        checkin_dir = os.path.join('therapy_data', 'checkins', patient_id)
        if os.path.exists(checkin_dir):
            shutil.rmtree(checkin_dir)

        # Log deletion
        log_activity('patient_deleted', {
            'patient_id': patient_id,
            'deleted_by': request.therapist['email']
        })

        return jsonify({
            'success': True,
            'message': 'All patient data deleted successfully'
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/privacy/export-patient-data/<patient_id>', methods=['GET'])
@mock_auth  # Using mock_auth instead of require_auth for development
def export_patient_data(patient_id):
    """GDPR compliance - export all patient data"""
    try:
        # Verify ownership
        patient_file = os.path.join('therapy_data', 'patients', f'patient_{patient_id}.json')
        if os.path.exists(patient_file):
            with open(patient_file, 'r') as f:
                patient = json.load(f)

            if patient.get('enrolledBy') != request.therapist['email'] and request.therapist['email'] != 'admin@system':
                return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        else:
            return jsonify({'success': False, 'error': 'Patient not found'}), 404

        # Collect all patient data
        export_data = {
            'patient_info': patient,
            'checkins': [],
            'export_date': datetime.now().isoformat(),
            'exported_by': request.therapist['email']
        }

        # Get all check-ins
        checkin_dir = os.path.join('therapy_data', 'checkins', patient_id)
        if os.path.exists(checkin_dir):
            for filename in sorted(os.listdir(checkin_dir)):
                if filename.endswith('.json'):
                    with open(os.path.join(checkin_dir, filename), 'r') as f:
                        export_data['checkins'].append(json.load(f))

        # Log export
        log_activity('patient_data_exported', {
            'patient_id': patient_id,
            'exported_by': request.therapist['email']
        })

        # Return as JSON file
        return Response(
            json.dumps(export_data, indent=2),
            mimetype='application/json',
            headers={
                'Content-Disposition': f'attachment; filename=patient_data_{patient_id}_{datetime.now().strftime("%Y%m%d")}.json'
            }
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============= HEALTH CHECK ENDPOINTS =============

@app.route('/api/health', methods=['GET'])
def health_check():
    """Simple health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'service': 'Enhanced Therapeutic Companion Backend',
        'version': '2.0',
        'features': [
            'Multi-user support with authentication',
            'Patient enrollment and management',
            'Daily check-ins (emotional, medication, physical)',
            '7-day weekly tracking',
            'Excel report generation',
            'Email report with system account',
            'Rate limiting for security',
            'GDPR compliance features',
            'Activity logging'
        ],
        'security': {
            'authentication': 'Token-based',
            'rate_limiting': 'Enabled',
            'cors': 'Configured',
            'https_only_cookies': os.environ.get('PRODUCTION', False)
        },
        'timestamp': datetime.now().isoformat()
    })


@app.route('/api/stats', methods=['GET'])
@mock_auth  # Using mock_auth instead of require_auth for development
def get_system_stats():
    """Get system statistics (admin only)"""
    if request.therapist['email'] != 'admin@system':
        return jsonify({'error': 'Admin access required'}), 403

    try:
        stats = {
            'therapists': 0,
            'patients': 0,
            'checkins': 0,
            'reports_generated': 0
        }

        # Count therapists
        therapists_dir = os.path.join('therapy_data', 'therapists')
        if os.path.exists(therapists_dir):
            stats['therapists'] = len([f for f in os.listdir(therapists_dir) if f.endswith('.json')])

        # Count patients
        patients_dir = os.path.join('therapy_data', 'patients')
        if os.path.exists(patients_dir):
            stats['patients'] = len([f for f in os.listdir(patients_dir) if f.endswith('.json')])

        # Count checkins
        checkins_dir = os.path.join('therapy_data', 'checkins')
        if os.path.exists(checkins_dir):
            for patient_dir in os.listdir(checkins_dir):
                patient_checkins = os.path.join(checkins_dir, patient_dir)
                if os.path.isdir(patient_checkins):
                    stats['checkins'] += len([f for f in os.listdir(patient_checkins) if f.endswith('.json')])

        # Count reports
        reports_dir = os.path.join('therapy_data', 'excel_exports')
        if os.path.exists(reports_dir):
            stats['reports_generated'] = len([f for f in os.listdir(reports_dir) if f.endswith('.xlsx')])

        return jsonify({
            'success': True,
            'stats': stats,
            'timestamp': datetime.now().isoformat()
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============= ERROR HANDLERS =============

@app.errorhandler(429)
def ratelimit_handler(e):
    return jsonify({
        'error': 'Rate limit exceeded',
        'message': str(e.description)
    }), 429


@app.errorhandler(500)
def internal_error(error):
    return jsonify({
        'error': 'Internal server error',
        'message': 'An unexpected error occurred'
    }), 500


if __name__ == "__main__":
    print("=" * 80)
    print("🏥 ENHANCED THERAPEUTIC COMPANION SYSTEM - DEVELOPMENT MODE")
    print("=" * 80)
    print("✅ Multi-user support with mock authentication")
    print("✅ Patient enrollment and management")
    print("✅ Daily check-ins (emotional, medication, physical)")
    print("✅ 7-day weekly tracking")
    print("✅ Excel report generation")
    print("✅ Email report with system account")
    print("✅ Rate limiting for security")
    print("✅ Activity logging")
    print("✅ GDPR compliance features")
    print("=" * 80)
    print("\n📁 Data storage locations:")
    print("  - therapy_data/therapists/ (therapist accounts)")
    print("  - therapy_data/patients/ (patient profiles)")
    print("  - therapy_data/checkins/ (daily check-ins)")
    print("  - therapy_data/excel_exports/ (generated reports)")
    print("  - therapy_data/logs/ (activity logs)")
    print("\n🔐 Security features:")
    print("  - Mock authentication enabled for development")
    print("  - Rate limiting on sensitive endpoints")
    print("  - Therapist-specific data access")
    print("\n🌐 Server running at: http://localhost:5000")
    print("=" * 80)

    # Run in production mode if environment variable is set
    if os.environ.get('PRODUCTION'):
        app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
    else:
        app.run(host='127.0.0.1', port=5000, debug=True)